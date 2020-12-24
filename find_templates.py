import psycopg2
from openpyxl import load_workbook


def check(templ, addr, full_adr):
    try:
        a = templ.format(type=addr[0], street=addr[1], type_house=addr[2], house=addr[3], type_apart=addr[4],
                         apart=addr[5])
        return a == full_adr
    except Exception:
        pass
    try:
        a = templ.format(type=addr[0], street=addr[1], type_house=addr[2], house=addr[3])
        return a == full_adr
    except Exception:
        pass

    try:
        a = templ.format(type=addr[0], street1=addr[1], street2=addr[2], type_house=addr[3], house=addr[4])
        return a == full_adr
    except Exception:
        pass

    try:
        a = templ.format(type=addr[0], street1=addr[1], street2=addr[2], type_house=addr[3], house=addr[4],
                         type_apart=addr[5], apart=addr[6])
        return a == full_adr
    except Exception:
        pass
    return False


class Finder():
    def __init__(self, path):
        self.wb = load_workbook(path)
        self.s = self.wb.active
        self.conn = psycopg2.connect(dbname='reimport2', user='cuba',
                                     password='cuba', host='localhost')
        self.cursor = self.conn.cursor()
        for i in range(1, 40000000):
            if self.s.cell(row=i, column=1).value == None:
                self.cnt = i
                break
        self.templ2 = {
            '{type} {street} {type_house} {house} {type_apart} {apart}': """
                select * from reimport_rtneo_refactor where street like '{street}%' 
                and house like '{house}|%' 
                and apartment like '{apart}'
            """,
            '{type} {street} {type_house} {house}': '',
            '{type} {street1} {street2} {type_house} {house}': '',
            '{type} {street1} {street2} {type_house} {house} {type_apart} {apart}': ''
        }

    def run(self):
        x = 0
        one, bad, few = 0, 0, 0
        for i in range(1, self.cnt):
            addr = self.s.cell(row=i, column=1).value
            addr = addr.split(' ')[4:]
            for templ in self.templ2.keys():
                if check(templ, addr, ' '.join(addr)):
                    if self.templ2[templ] != '':
                        self.cursor.execute(self.templ2[templ].format(type=addr[0], street=addr[1].upper(), type_house=addr[2], house=addr[3].replace(',', ''), type_apart=addr[4],
                         apart=addr[5]))

                        records = self.cursor.fetchall()
                        if len(records) > 1:
                            few += 1
                        elif len(records) == 1:
                            one += 1
                        else:
                            bad += 1

                        print(one, bad, few)
                    break
            else:
                print(addr)
                x += 1
