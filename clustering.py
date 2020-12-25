import pandas as pd
import numpy as np
from scipy.cluster.hierarchy import linkage, fcluster, dendrogram
from openpyxl import load_workbook
import matplotlib.pyplot as plt


class Clustering:
    def __init__(self, name):
        self.wb = load_workbook(name)

    def read_data(self):
        s = self.wb.active
        data = []
        for i in range(1, 400000):
            address = s.cell(row=i, column=1).value
            if address == None:
                break
            ord_address = [ord(x) for x in address]
            while len(ord_address) < 62:
                ord_address.append(0)
            data.append(ord_address)
        # print(data)
        self.data = data

    def clustering(self):
        X = pd.DataFrame(self.data)
        Z = linkage(X, method='average')
        print(Z)
        plt.figure(figsize=(25, 10))
        plt.title('Hierarchical Clustering Dendrogram')
        plt.xlabel('sample index')
        plt.ylabel('distance')
        dendrogram(Z, orientation='left', color_threshold=0.0)
        plt.show()

        label = fcluster(Z, 2500, criterion='distance')

        print(np.unique(label))

        X.loc[:, 'label'] = label
        f = open('clustering.txt', 'w')
        for i, group in X.groupby('label'):
            f.write('=' * 20)
            f.write(' cluster {} '.format(i))
            f.write('=' * 20)
            f.write('\n')
            x = group.values
            # print(x)
            for j in x:
                for jj in j[:-1]:
                    if jj > 0: f.write(chr(jj))
                f.write('\n')
            f.write('\n')